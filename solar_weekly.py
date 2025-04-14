import streamlit as st
import pandas as pd
import io
import openpyxl
from datetime import date, timedelta

def solar_weekly_tab():
    st.title("Solar Weekly Report (UNDER UDVIKLING)")
    
    # Beregn forrige uge (fx den uge, hvor rapporten skal laves)
    today = date.today()
    last_week_date = today - timedelta(days=7)
    week_number = last_week_date.isocalendar()[1]
    week_str = f"Uge {week_number:02d}"
    
    st.markdown(f"### Rapport for {week_str}")

    # Upload datafilen med ruteinfo
    st.markdown("Upload datafil (Excel)")
    uploaded_file = st.file_uploader("Upload Data_routestats Excel-fil", type=["xlsx"], key="data_routestats")
    
    if uploaded_file:
        try:
            # Læs den uploadede fil (forventet med standard headers)
            df = pd.read_excel(uploaded_file, engine="openpyxl")
            
            # Tjek at kolonnen "STATUS" findes
            if "STATUS" not in df.columns:
                st.error("Kolonnen 'STATUS' findes ikke i den uploadede fil.")
                return
            
            # Bevar kun rækker, hvor STATUS er "Completed" (case-insensitive)
            df = df[df["STATUS"].astype(str).str.strip().str.lower() == "completed"]
            if df.empty:
                st.error("Ingen rækker med STATUS 'Completed' fundet.")
                return
            
            # Forvent, at kolonner B, C og D skal formateres til numre.
            # Hvis vi antager, at den første række med header er på række 0 og:
            # - Kolonne B: df.columns[1]
            # - Kolonne C: df.columns[2]
            # - Kolonne D: df.columns[3] med header "RUTE REFERENCE"
            try:
                col_B = df.columns[1]
                col_C = df.columns[2]
                col_D = df.columns[3]  # Forventer "RUTE REFERENCE"
            except Exception as e:
                st.error("Datafilen har ikke mindst 4 kolonner, så konvertering af B, C og D kunne ikke udføres.")
                return
            
            df[col_B] = pd.to_numeric(df[col_B], errors="coerce")
            df[col_C] = pd.to_numeric(df[col_C], errors="coerce")
            df[col_D] = pd.to_numeric(df[col_D], errors="coerce")
            
            # Filtrer på RUTE REFERENCE - behold kun de tal der starter med "80"
            df = df[df[col_D].astype(str).str.startswith("80")]
            if df.empty:
                st.error("Ingen rækker med RUTE REFERENCE der starter med '80' fundet.")
                return
            
            # Sorter på RUTE REFERENCE (fra mindst til størst)
            df = df.sort_values(by=col_D)
            
            # Udtræk unikke værdier fra kolonne D ("RUTE REFERENCE")
            unique_refs = df[col_D].drop_duplicates().tolist()
            
            st.success("Data er filtreret og sorteret korrekt.")
            st.write("Unikke RUTE REFERENCE værdier:", unique_refs)
            
            # Åbn Excel-template-filen med den ønskede formatering og formler (f.eks. "TEMPLATE SHORT - Solar weekly report uge XX (formler).xlsx")
            template_path = "TEMPLATE SHORT - Solar weekly report uge XX (formler).xlsx"
            wb = openpyxl.load_workbook(template_path)
            
            # Arbejd i fanen "Uge xx" – tjek at den findes
            if "Uge xx" not in wb.sheetnames:
                st.error("Template-filen mangler fanen 'Uge xx'.")
                return
            ws_uge = wb["Uge xx"]
            
            # Ryd eksisterende data fra række 4 og ned (bevar rækker 1-3 med header og evt. formler i række 4)
            max_row = ws_uge.max_row
            for row in ws_uge.iter_rows(min_row=4, max_row=max_row):
                for cell in row:
                    cell.value = None
            
            # Indsæt de unikke booking referencer fra unique_refs i kolonne A ("Booking ref.") fra række 4 og ned
            start_row = 4
            for i, ref in enumerate(unique_refs, start=start_row):
                ws_uge.cell(row=i, column=1, value=ref)
            
            # Kopier XLOOKUP-formlerne fra række 4, kolonner B til L, og tilpas dem til de nye rækker
            # Forudsætning: Template-filen indeholder de korrekte formler i række 4 i fanen "Uge xx"
            formulas = {}
            for col in range(2, 13):  # Kolonner B (2) til L (12)
                cell = ws_uge.cell(row=start_row, column=col)
                if cell.data_type == 'f':  # hvis cellen indeholder en formel
                    formulas[col] = cell.value
            
            num_rows = len(unique_refs)
            # For hver række under række 4 indsætter vi en tilsvarende formel til de øvrige kolonner
            for row in range(start_row, start_row + num_rows):
                for col in range(2, 13):
                    # Hvis cellen er tom, indsæt formlen (juster evt. rækkehenvisningen)
                    if ws_uge.cell(row=row, column=col).value is None:
                        if col in formulas:
                            original_formula = formulas[col]
                            # Eksempel: erstat "$A4" med "$A{row}" – dette er en simpel udskiftning og forudsætter, at formlen benytter denne notation.
                            new_formula = original_formula.replace("$A4", f"$A{row}")
                            ws_uge.cell(row=row, column=col, value=new_formula)
            
            # Fjern alle andre faner, så kun fanen "Uge xx" er tilbage i den endelige fil
            for sheet in wb.sheetnames:
                if sheet != "Uge xx":
                    del wb[sheet]
            
            # Gem den redigerede workbook til en buffer og gør den klar til download
            towrite = io.BytesIO()
            wb.save(towrite)
            towrite.seek(0)
            
            st.success("Rapporten er genereret med den ønskede formatering.")
            st.download_button(
                label=f"Download rapport for {week_str}",
                data=towrite,
                file_name=f"solar_weekly_report_{week_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"Fejl under behandling af filen: {e}")
